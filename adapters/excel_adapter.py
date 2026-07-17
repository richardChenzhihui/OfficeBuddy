"""Excel adapter wrapping openpyxl"""
from typing import Any, Optional, List, Tuple, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart
from openpyxl.utils import get_column_letter

from ..schemas.operations import StyleParams, FillMode
from ..schemas.selector import ExcelSelector


class ExcelAdapter:
    """Adapter for openpyxl operations"""
    
    @staticmethod
    def read_cells(worksheet: Any, coords: Optional[Tuple[int, int, int, int]], include_formula: bool = False) -> Dict[str, Any]:
        """Read cells from worksheet"""
        if coords is None:
            # Read all used cells
            data = []
            formulas = []
            for row in worksheet.iter_rows(values_only=False):
                row_data = []
                row_formulas = []
                for cell in row:
                    row_data.append(cell.value)
                    if include_formula and cell.data_type == 'f':
                        row_formulas.append(cell.value)
                    else:
                        row_formulas.append(None)
                data.append(row_data)
                formulas.append(row_formulas)
        else:
            min_row, min_col, max_row, max_col = coords
            data = []
            formulas = []
            for row_idx in range(min_row, (max_row or worksheet.max_row) + 1):
                row_data = []
                row_formulas = []
                for col_idx in range(min_col, (max_col or worksheet.max_column) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    row_data.append(cell.value)
                    if include_formula and cell.data_type == 'f':
                        row_formulas.append(cell.value)
                    else:
                        row_formulas.append(None)
                data.append(row_data)
                formulas.append(row_formulas)
        
        return {
            "data": data,
            "formulas": formulas if include_formula else None
        }
    
    @staticmethod
    def write_cells(worksheet: Any, coords: Optional[Tuple[int, int, int, int]], 
                   values: List[List[Any]], fill_mode: FillMode = FillMode.OVERWRITE) -> Dict[str, Any]:
        """Write cells to worksheet"""
        if coords is None:
            # Write starting from A1
            min_row, min_col = 1, 1
            max_row = min_row + len(values) - 1
            max_col = min_col + (max(len(row) for row in values) if values else 0) - 1
        else:
            min_row, min_col, max_row, max_col = coords
            if max_row is None:
                max_row = min_row + len(values) - 1
            if max_col is None:
                max_col = min_col + (max(len(row) for row in values) if values else 0) - 1
        
        affected = []
        for i, row_values in enumerate(values):
            row_idx = min_row + i
            for j, value in enumerate(row_values):
                col_idx = min_col + j
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if fill_mode == FillMode.OVERWRITE:
                    cell.value = value
                elif fill_mode == FillMode.APPEND:
                    if cell.value is None:
                        cell.value = value
                    else:
                        cell.value = str(cell.value) + str(value)
                elif fill_mode == FillMode.MERGE:
                    if cell.value is None:
                        cell.value = value
                
                affected.append(f"{get_column_letter(col_idx)}{row_idx}")
        
        return {
            "affected": affected,
            "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        }
    
    @staticmethod
    def edit_formula(worksheet: Any, coords: Optional[Tuple[int, int, int, int]], formula: str) -> Dict[str, Any]:
        """Edit formula in cells"""
        if coords is None:
            return {"affected": [], "error": "Coordinates required for formula editing"}
        
        min_row, min_col, max_row, max_col = coords
        affected = []
        
        for row_idx in range(min_row, (max_row or min_row) + 1):
            for col_idx in range(min_col, (max_col or min_col) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = formula
                affected.append(f"{get_column_letter(col_idx)}{row_idx}")
        
        return {
            "affected": affected,
            "formula": formula
        }
    
    @staticmethod
    def apply_style(worksheet: Any, coords: Optional[Tuple[int, int, int, int]], 
                   style: StyleParams) -> Dict[str, Any]:
        """Apply style to cells"""
        if coords is None:
            return {"affected": [], "error": "Coordinates required for style editing"}
        
        min_row, min_col, max_row, max_col = coords
        affected = []
        
        # Build font
        font_kwargs = {}
        if style.font_name:
            font_kwargs["name"] = style.font_name
        if style.font_size:
            font_kwargs["size"] = style.font_size
        if style.bold is not None:
            font_kwargs["bold"] = style.bold
        if style.italic is not None:
            font_kwargs["italic"] = style.italic
        if style.underline is not None:
            font_kwargs["underline"] = "single" if style.underline else None
        
        font = Font(**font_kwargs) if font_kwargs else None
        
        # Build fill
        fill = None
        if style.bg_color:
            fill = PatternFill(start_color=style.bg_color, end_color=style.bg_color, fill_type="solid")
        
        # Build alignment
        alignment = None
        if style.alignment:
            alignment = Alignment(horizontal=style.alignment)
        
        # Apply to cells
        for row_idx in range(min_row, (max_row or worksheet.max_row) + 1):
            for col_idx in range(min_col, (max_col or worksheet.max_column) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
                if style.number_format:
                    cell.number_format = style.number_format
                
                affected.append(f"{get_column_letter(col_idx)}{row_idx}")
        
        return {
            "affected": affected,
            "style_applied": {
                "font": bool(font),
                "fill": bool(fill),
                "alignment": bool(alignment),
                "number_format": bool(style.number_format)
            }
        }
    
    @staticmethod
    def insert_rows_cols(worksheet: Any, position: int, count: int, 
                        insert_type: str = "row") -> Dict[str, Any]:
        """Insert rows or columns"""
        if insert_type == "row":
            worksheet.insert_rows(position, count)
            affected = [f"Row {position} to {position + count - 1}"]
        else:  # column
            worksheet.insert_cols(position, count)
            col_letter = get_column_letter(position)
            affected = [f"Column {col_letter}"]
        
        return {
            "affected": affected,
            "type": insert_type,
            "position": position,
            "count": count
        }
    
    @staticmethod
    def create_chart(worksheet: Any, data_range: str, chart_type: str, 
                    options: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Create chart in worksheet"""
        options = options or {}
        
        # Parse data range
        from openpyxl.utils import range_boundaries
        try:
            min_col, min_row, max_col, max_row = range_boundaries(data_range)
        except:
            return {"error": f"Invalid data range: {data_range}"}
        
        # Create chart based on type
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            return {"error": f"Unsupported chart type: {chart_type}"}
        
        # Set data
        from openpyxl.chart.reference import Reference
        data = Reference(worksheet, min_col=min_col, min_row=min_row, 
                        max_col=max_col, max_row=max_row)
        chart.add_data(data, titles_from_data=options.get("titles_from_data", True))
        
        # Set title
        if "title" in options:
            chart.title = options["title"]
        
        # Add chart to worksheet
        chart_cell = options.get("chart_cell", "E1")
        worksheet.add_chart(chart, chart_cell)
        
        return {
            "chart_type": chart_type,
            "data_range": data_range,
            "chart_cell": chart_cell
        }
    
    @staticmethod
    def conditional_select(worksheet: Any, condition: Dict[str, Any]) -> List[str]:
        """Select cells based on condition"""
        matches = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                match = True
                
                if "contains" in condition:
                    if cell.value is None or condition["contains"] not in str(cell.value):
                        match = False
                
                if "value_gt" in condition and match:
                    try:
                        if cell.value is None or float(cell.value) <= condition["value_gt"]:
                            match = False
                    except (ValueError, TypeError):
                        match = False
                
                if "has_formula" in condition and match:
                    if (condition["has_formula"] and cell.data_type != 'f') or \
                       (not condition["has_formula"] and cell.data_type == 'f'):
                        match = False
                
                if match:
                    matches.append(f"{get_column_letter(cell.column)}{cell.row}")
        
        return matches
