"""Regression tests for bench/BUGS.md OA-1 … OA-6.

Everything here is offline: no Word/Excel automation, no rendering. The parts
of each bug that real Excel exposed as "-50 参数错误" are checked at the level
they are actually caused — the OOXML package on disk.
"""
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from office_agent.adapters.excel_adapter import ExcelAdapter
from office_agent.core.session import (
    EditSession,
    _assert_extension_matches_content_type,
    _has_vba,
)
from office_agent.tools import REGISTRY, ToolContext

MACRO_CT = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"


def _plain_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "销售明细"
    ws.append(["日期", "地区", "产品", "数量", "金额"])
    for i in range(1, 13):
        ws.append([f"2026-01-{i:02d}", "华东", "A", i, i * 100])
    wb.save(str(path))
    return path


def _package(path: Path):
    with zipfile.ZipFile(path) as zf:
        names = set(zf.namelist())
        cts = zf.read("[Content_Types].xml").decode()
        rels = zf.read("xl/_rels/workbook.xml.rels").decode()
    return names, cts, rels


# --------------------------------------------------------------------- OA-1


def test_plain_xlsx_roundtrip_is_not_a_macro_package(tmp_path):
    """keep_vba on a VBA-free workbook used to flip the content type and add a
    dangling vbaProject relationship — real Excel then refuses to open it."""
    src = _plain_workbook(tmp_path / "plain.xlsx")
    session = EditSession(str(src))
    try:
        session.doc["销售明细"]["A1"] = "日期(改)"
        out = session.save_to(str(tmp_path / "out.xlsx"))
    finally:
        session.cleanup()

    names, cts, rels = _package(out)
    assert MACRO_CT not in cts, "plain .xlsx must not carry the macro content type"
    assert "vbaProject" not in rels, "no vbaProject relationship without a VBA project"
    assert "xl/vbaProject.bin" not in names
    assert load_workbook(str(out))["销售明细"]["A1"].value == "日期(改)"


def test_has_vba_detects_only_real_macro_packages(tmp_path):
    src = _plain_workbook(tmp_path / "plain.xlsx")
    assert _has_vba(src) is False

    macro = tmp_path / "macro.xlsm"
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(macro, "w") as zout:
        for item in zin.infolist():
            zout.writestr(item, zin.read(item.filename))
        zout.writestr("xl/vbaProject.bin", b"\x00fake-vba")
    assert _has_vba(macro) is True


def test_xlsm_still_loads_with_keep_vba(tmp_path):
    """A real macro workbook must keep its VBA project through an edit."""
    src = _plain_workbook(tmp_path / "plain.xlsx")
    macro = tmp_path / "macro.xlsm"
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(macro, "w") as zout:
        for item in zin.infolist():
            zout.writestr(item, zin.read(item.filename))
        zout.writestr("xl/vbaProject.bin", b"\x00fake-vba")

    session = EditSession(str(macro))
    try:
        assert session.doc.vba_archive is not None
        session.doc["销售明细"]["A1"] = "x"
        out = session.save_to(str(tmp_path / "out.xlsm"), accept_fidelity_loss=True)
    finally:
        session.cleanup()
    names, cts, _ = _package(out)
    assert "xl/vbaProject.bin" in names
    assert MACRO_CT in cts  # correct for .xlsm


def test_content_type_guard_rejects_macro_ct_on_xlsx(tmp_path):
    """The belt-and-braces assertion that catches any future regression."""
    src = _plain_workbook(tmp_path / "plain.xlsx")
    _assert_extension_matches_content_type(src)  # healthy file passes

    broken = tmp_path / "broken.xlsx"
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(broken, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "[Content_Types].xml":
                data = data.replace(
                    b"application/vnd.openxmlformats-officedocument."
                    b"spreadsheetml.sheet.main+xml",
                    MACRO_CT.encode(),
                )
            zout.writestr(item, data)
    with pytest.raises(RuntimeError, match="macro-workbook content type"):
        _assert_extension_matches_content_type(broken)


def test_content_type_guard_rejects_dangling_vba_relationship(tmp_path):
    src = _plain_workbook(tmp_path / "plain.xlsm")
    dangling = tmp_path / "dangling.xlsm"
    rel = (
        '<Relationship Id="rIdVba" Type="http://schemas.microsoft.com/office/2006/'
        'relationships/vbaProject" Target="vbaProject.bin"/>'
    )
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(dangling, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/_rels/workbook.xml.rels":
                data = data.replace(b"</Relationships>", rel.encode() + b"</Relationships>")
            zout.writestr(item, data)
    with pytest.raises(RuntimeError, match="no xl/vbaProject.bin"):
        _assert_extension_matches_content_type(dangling)


# --------------------------------------------------------------------- OA-2


def test_freeze_panes_tool(tmp_path):
    src = _plain_workbook(tmp_path / "f.xlsx")
    ctx = ToolContext()
    doc_id = REGISTRY.dispatch(ctx, "open_document", {"file_path": str(src)})["doc_id"]
    try:
        res = REGISTRY.dispatch(
            ctx, "excel_freeze_panes", {"doc_id": doc_id, "sheet": "销售明细", "cell": "A2"}
        )
        assert res["success"] and res["freeze_panes"] == "A2"
        assert ctx.sessions.get(doc_id).doc["销售明细"].freeze_panes == "A2"

        structure = REGISTRY.dispatch(ctx, "get_structure", {"doc_id": doc_id})
        assert structure["sheets"][0]["freeze_panes"] == "A2"

        # Unfreeze via null, and via the A1 no-op form.
        assert REGISTRY.dispatch(
            ctx, "excel_freeze_panes", {"doc_id": doc_id, "sheet": "销售明细", "cell": None}
        )["freeze_panes"] is None
        assert ctx.sessions.get(doc_id).doc["销售明细"].freeze_panes is None
    finally:
        ctx.sessions.close_all()


def test_freeze_panes_rejects_a_range(tmp_path):
    src = _plain_workbook(tmp_path / "f.xlsx")
    ctx = ToolContext()
    doc_id = REGISTRY.dispatch(ctx, "open_document", {"file_path": str(src)})["doc_id"]
    try:
        res = REGISTRY.dispatch(
            ctx,
            "excel_freeze_panes",
            {"doc_id": doc_id, "sheet": "销售明细", "cell": "A2:B5"},
        )
        assert not res["success"] and "ONE cell" in res["error"]
    finally:
        ctx.sessions.close_all()


def test_freeze_panes_survives_save(tmp_path):
    src = _plain_workbook(tmp_path / "f.xlsx")
    session = EditSession(str(src))
    try:
        ExcelAdapter.freeze_panes(session.doc["销售明细"], "B2")
        out = session.save_to(str(tmp_path / "frozen.xlsx"))
    finally:
        session.cleanup()
    assert load_workbook(str(out))["销售明细"].freeze_panes == "B2"


# --------------------------------------------------------------------- OA-3


def test_sheet_lifecycle_tool(tmp_path):
    src = _plain_workbook(tmp_path / "s.xlsx")
    ctx = ToolContext()
    doc_id = REGISTRY.dispatch(ctx, "open_document", {"file_path": str(src)})["doc_id"]
    try:
        created = REGISTRY.dispatch(
            ctx,
            "excel_manage_sheet",
            {"doc_id": doc_id, "action": "create", "sheet": "汇总"},
        )
        assert created["success"] and created["sheet"] == "汇总"

        # A cross-sheet formula into the new sheet — the E4 scenario.
        assert REGISTRY.dispatch(
            ctx,
            "excel_edit_formula",
            {
                "doc_id": doc_id,
                "sheet": "汇总",
                "range": "B2",
                "formula": "=SUMIF('销售明细'!B:B,A2,'销售明细'!E:E)",
            },
        )["success"]

        renamed = REGISTRY.dispatch(
            ctx,
            "excel_manage_sheet",
            {"doc_id": doc_id, "action": "rename", "sheet": "汇总", "new_name": "地区汇总"},
        )
        assert renamed["success"] and "地区汇总" in renamed["sheets"]

        copied = REGISTRY.dispatch(
            ctx,
            "excel_manage_sheet",
            {"doc_id": doc_id, "action": "copy", "sheet": "地区汇总", "new_name": "备份"},
        )
        assert copied["success"] and copied["sheet"] == "备份"

        deleted = REGISTRY.dispatch(
            ctx, "excel_manage_sheet", {"doc_id": doc_id, "action": "delete", "sheet": "备份"}
        )
        assert deleted["success"] and "备份" not in deleted["sheets"]

        session = ctx.sessions.get(doc_id)
        out = session.save_to(str(tmp_path / "out.xlsx"))
        wb = load_workbook(str(out))
        assert wb.sheetnames == ["销售明细", "地区汇总"]
        assert wb["地区汇总"]["B2"].value.startswith("=SUMIF")
    finally:
        ctx.sessions.close_all()


@pytest.mark.parametrize(
    "payload,expect",
    [
        ({"action": "create", "sheet": "销售明细"}, "already exists"),
        ({"action": "create", "sheet": "a" * 32}, "at most 31"),
        ({"action": "create", "sheet": "a/b"}, "forbids"),
        ({"action": "delete", "sheet": "没有这个表"}, "not found"),
        ({"action": "rename", "sheet": "销售明细"}, "requires new_name"),
        ({"action": "explode", "sheet": "销售明细"}, "Unknown action"),
        ({"action": "create"}, "'sheet' is required"),
    ],
)
def test_sheet_tool_rejects_bad_input(tmp_path, payload, expect):
    src = _plain_workbook(tmp_path / "s.xlsx")
    ctx = ToolContext()
    doc_id = REGISTRY.dispatch(ctx, "open_document", {"file_path": str(src)})["doc_id"]
    try:
        res = REGISTRY.dispatch(ctx, "excel_manage_sheet", {"doc_id": doc_id, **payload})
        assert not res["success"] and expect in res["error"]
    finally:
        ctx.sessions.close_all()


def test_cannot_delete_the_last_sheet(tmp_path):
    path = tmp_path / "one.xlsx"
    Workbook().save(str(path))
    ctx = ToolContext()
    doc_id = REGISTRY.dispatch(ctx, "open_document", {"file_path": str(path)})["doc_id"]
    try:
        res = REGISTRY.dispatch(
            ctx, "excel_manage_sheet", {"doc_id": doc_id, "action": "delete", "sheet": "Sheet"}
        )
        assert not res["success"] and "at least one sheet" in res["error"]
    finally:
        ctx.sessions.close_all()


# --------------------------------------------------------------------- OA-5


def test_default_chart_anchor_clears_the_data(tmp_path):
    """Data occupies A1:E13; the default anchor must not land inside it."""
    src = _plain_workbook(tmp_path / "c.xlsx")
    session = EditSession(str(src))
    try:
        ws = session.doc["销售明细"]
        res = ExcelAdapter.create_chart(ws, "A1:B13", "bar", {})
        assert res["chart_cell"] == "G1"  # max_column(E=5) + 2
        assert ExcelAdapter.create_chart(ws, "A1:B13", "bar", {"chart_cell": "B20"})[
            "chart_cell"
        ] == "B20"
    finally:
        session.cleanup()


def test_default_chart_anchor_on_an_empty_sheet(tmp_path):
    path = tmp_path / "e.xlsx"
    wb = Workbook()
    wb.active.append(["x", 1])
    wb.save(str(path))
    session = EditSession(str(path))
    try:
        assert ExcelAdapter.create_chart(session.doc.active, "A1:B1", "bar", {})[
            "chart_cell"
        ] == "D1"
    finally:
        session.cleanup()
