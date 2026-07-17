"""Part-inventory fidelity guard tests (design: docs/edit-layer-designs/excel-fidelity-guard.md)."""
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _xlsx_fixtures import inject_foreign_parts

from office_agent.core.session import (
    EditSession,
    FidelityLossError,
    _categorize,
    _zip_inventory,
)
from office_agent.tools import REGISTRY, ToolContext


def _rich_workbook(path: Path) -> None:
    """Every openpyxl-modeled feature: must round-trip losslessly."""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    for row in [["h1", "h2"], ["a", 1], ["b", 2]]:
        ws.append(row)
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=3))
    ws.add_chart(chart, "E1")
    ws.conditional_formatting.add(
        "B2:B3",
        CellIsRule(
            operator="greaterThan",
            formula=["1"],
            fill=PatternFill("solid", start_color="FF0000", end_color="FF0000"),
        ),
    )
    dv = DataValidation(type="list", formula1='"x,y"')
    ws.add_data_validation(dv)
    dv.add("D1")
    table = Table(displayName="T1", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.merge_cells("D3:E3")
    wb.save(str(path))


def test_zip_inventory_path_vs_bytes_parity(tmp_path):
    p = tmp_path / "x.xlsx"
    Workbook().save(str(p))
    assert _zip_inventory(p) == _zip_inventory(p.read_bytes())


def test_categorize_prefixes():
    assert _categorize("xl/threadedComments/tc1.xml") == "threaded comment"
    assert _categorize("xl/charts/chart1.xml") == "chart"
    assert _categorize("weird/part.bin") == "unrecognized part"


def test_no_false_positive_on_openpyxl_feature_surface(tmp_path):
    src = tmp_path / "rich.xlsx"
    _rich_workbook(src)
    session = EditSession(str(src))
    try:
        assert session.fidelity_report == [], session.fidelity_report
        out = session.save_to(str(tmp_path / "out.xlsx"))  # no consent flag needed
        assert out.exists()
    finally:
        session.cleanup()


def test_true_positive_names_categories_and_blocks_save(tmp_path):
    src = tmp_path / "foreign.xlsx"
    Workbook().save(str(src))
    injected = inject_foreign_parts(
        src, ["threaded_comment", "person", "slicer", "custom_xml"]
    )
    ctx = ToolContext()
    try:
        opened = REGISTRY.dispatch(ctx, "open_document", {"file_path": str(src)})
        assert opened["success"], opened
        assert "warnings" in opened
        for name in injected:
            assert name in opened["warnings"][0]
        doc_id = opened["doc_id"]

        report = REGISTRY.dispatch(ctx, "excel_fidelity_report", {"doc_id": doc_id})
        categories = {e["category"] for e in report["at_risk_parts"]}
        assert {"threaded comment", "comment-author registry", "slicer",
                "custom XML part"} <= categories

        blocked = REGISTRY.dispatch(
            ctx, "save_document", {"doc_id": doc_id, "path": str(tmp_path / "o.xlsx")}
        )
        assert blocked["success"] is False
        assert blocked["error_type"] == "FidelityLossError"
        assert "threadedComment" in blocked["error"]

        consented = REGISTRY.dispatch(
            ctx,
            "save_document",
            {
                "doc_id": doc_id,
                "path": str(tmp_path / "o.xlsx"),
                "accept_fidelity_loss": True,
            },
        )
        assert consented["success"], consented
        saved_names = set(_zip_inventory(Path(consented["saved_path"])))
        for name in injected:  # honest loss: dropped parts really are gone
            assert name not in saved_names
    finally:
        ctx.sessions.close_all()


def test_size_change_alone_never_flags(tmp_path):
    src = tmp_path / "chart.xlsx"
    _rich_workbook(src)  # chart present: sizes change on round trip
    session = EditSession(str(src))
    try:
        assert session.fidelity_report == []
    finally:
        session.cleanup()


def test_keep_vba_rescues_vba_part(tmp_path):
    import shutil
    import zipfile

    src = tmp_path / "macro.xlsm"
    Workbook().save(str(src))
    tmp = src.with_suffix(".t.xlsm")
    shutil.copy2(src, tmp)
    with zipfile.ZipFile(src) as a, zipfile.ZipFile(tmp, "w") as b:
        for item in a.infolist():
            b.writestr(item, a.read(item.filename))
        b.writestr("xl/vbaProject.bin", b"\x00fakevba")
    tmp.replace(src)
    session = EditSession(str(src))
    try:
        lost_names = [e.path for e in session.fidelity_report]
        assert "xl/vbaProject.bin" not in lost_names  # keep_vba=True rescued it
    finally:
        session.cleanup()


def test_rich_text_preserved_and_read_as_plain_string(tmp_path):
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    from office_agent.adapters.excel_adapter import ExcelAdapter

    src = tmp_path / "rich_text.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = CellRichText([TextBlock(InlineFont(b=True), "Bold"), " plain"])
    wb.save(str(src))

    session = EditSession(str(src))
    try:
        ws2 = session.doc.active
        assert isinstance(ws2["A1"].value, CellRichText)  # not flattened at load
        result = ExcelAdapter.read_cells(ws2, (1, 1, 1, 1))
        assert result["data"][0][0] == "Bold plain"  # JSON-safe plain string
        out = session.save_to(str(tmp_path / "rt_out.xlsx"))
        wb3 = load_workbook(str(out), rich_text=True)
        assert isinstance(wb3.active["A1"].value, CellRichText)  # survived save
    finally:
        session.cleanup()


def test_word_docs_unaffected_by_guard(word_doc_path):
    session = EditSession(str(word_doc_path))
    try:
        assert session.fidelity_report == []
        assert session.fidelity_warnings() == []
    finally:
        session.cleanup()
