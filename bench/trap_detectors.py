"""Mechanical detectors for the visual traps.

Scope discipline: everything in this file decides a *geometric or attribute
fact* — is a slot set, does this width exceed that width, do these two ranges
intersect, what is the contrast ratio of these two colors, which page is this
text block on. Nothing here decides whether a system *behaved well*; that is a
semantic judgement and belongs to the LLM judge in judge_visual.py.

Every detector returns:
    {"trap": <id>, "present": bool | None, "evidence": {...}}
`present=None` means the detector could not decide (e.g. a number format it
does not model) — the run is escalated to the judge rather than silently
scored as clean.
"""
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
XDR = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
ORel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

CJK = re.compile(r"[㐀-䶿一-鿿豈-﫿぀-ヿ가-힯]")


def _result(trap: str, present, **evidence) -> dict:
    return {"trap": trap, "present": present, "evidence": evidence}


# ==========================================================================
# VT1 — East-Asian font slot
# ==========================================================================
def _docx_part(path: Path, name: str) -> ET.Element | None:
    with zipfile.ZipFile(path) as z:
        if name not in z.namelist():
            return None
        return ET.fromstring(z.read(name))


def _rfonts_slots(rpr: ET.Element | None) -> dict:
    """Which font slots does this rPr explicitly pin?"""
    if rpr is None:
        return {}
    rf = rpr.find(f"{W}rFonts")
    if rf is None:
        return {}
    out = {}
    for slot in ("ascii", "hAnsi", "eastAsia", "cs"):
        v = rf.get(f"{W}{slot}")
        if v:
            out[slot] = v
    for slot in ("asciiTheme", "hAnsiTheme", "eastAsiaTheme"):
        v = rf.get(f"{W}{slot}")
        if v:
            out[slot] = v
    return out


def detect_cjk_slot(path: Path) -> dict:
    """A run holding CJK text that pins the Latin slots but leaves eastAsia
    unset forces Word to pick a fallback face per character at layout time.

    Nothing pinned at all is fine — the theme resolves it consistently. The
    defect is specifically the *asymmetric* pin.
    """
    doc = _docx_part(path, "word/document.xml")
    if doc is None:
        return _result("cjk_font_slot", None, reason="no word/document.xml")

    styles = _docx_part(path, "word/styles.xml")
    style_slots: dict[str, dict] = {}
    default_slots: dict = {}
    if styles is not None:
        dd = styles.find(f"{W}docDefaults/{W}rPrDefault/{W}rPr")
        default_slots = _rfonts_slots(dd)
        for st in styles.findall(f"{W}style"):
            sid = st.get(f"{W}styleId")
            if sid:
                style_slots[sid] = _rfonts_slots(st.find(f"{W}rPr"))

    offenders = []
    total_cjk_runs = 0
    for para in doc.iter(f"{W}p"):
        pstyle_el = para.find(f"{W}pPr/{W}pStyle")
        pstyle = pstyle_el.get(f"{W}val") if pstyle_el is not None else None
        para_rpr = _rfonts_slots(para.find(f"{W}pPr/{W}rPr"))
        for run in para.findall(f"{W}r"):
            text = "".join(t.text or "" for t in run.findall(f"{W}t"))
            if not CJK.search(text):
                continue
            total_cjk_runs += 1
            # Resolve slot by slot: run rPr > paragraph mark rPr > paragraph
            # style > docDefaults.
            chain = [
                _rfonts_slots(run.find(f"{W}rPr")),
                para_rpr,
                style_slots.get(pstyle or "", {}),
                default_slots,
            ]

            def resolve(*slots):
                """Returns (value, level_index) — which level won matters."""
                for i, level in enumerate(chain):
                    for s in slots:
                        if s in level:
                            return level[s], i
                return None, None

            latin, latin_lvl = resolve("ascii", "hAnsi", "asciiTheme", "hAnsiTheme")
            east, east_lvl = resolve("eastAsia", "eastAsiaTheme")
            # The defect is an *asymmetric* pin: this run explicitly chose a
            # Latin face, but its CJK glyphs still resolve from further down
            # the chain (usually the theme's minorEastAsia). The requested
            # font therefore never reaches the Chinese text. Both slots
            # resolving at the same level is fine — that is a consistent
            # document, whether the source is the run or docDefaults.
            if latin is not None and (east is None or east_lvl > latin_lvl):
                offenders.append(
                    {
                        "text": text[:40],
                        "latin": latin,
                        "latin_level": latin_lvl,
                        "eastasia": east,
                        "eastasia_level": east_lvl,
                    }
                )

    return _result(
        "cjk_font_slot",
        bool(offenders),
        cjk_runs=total_cjk_runs,
        offending_runs=len(offenders),
        samples=offenders[:5],
    )


# ==========================================================================
# VT2 — Excel column overflow (########)
# ==========================================================================
_FMT_TOKENS = re.compile(r'\[[^\]]*\]|"[^"]*"|\\.')


def _format_width(value, fmt: str):
    """Width in characters of `value` rendered under `fmt`.

    Models the number formats an agent actually reaches for. Returns None for
    anything outside that set so the caller can escalate rather than guess.
    """
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return None
    body = fmt.split(";")[0]
    literals = "".join(
        m.group(0).strip('"').lstrip("\\").strip("[]")
        for m in _FMT_TOKENS.finditer(body)
        if not m.group(0).startswith("[")
    )
    core = _FMT_TOKENS.sub("", body)
    prefix_suffix = len(literals)

    if not re.fullmatch(r"[#0,.%\s]*", core) or not re.search(r"[#0]", core):
        return None  # dates, scientific, fractions, text formats — not modelled

    scaled = value * 100 if "%" in core else value
    decimals = 0
    if "." in core:
        decimals = len(re.sub(r"[^0#]", "", core.split(".", 1)[1]))
    grouped = "," in core.split(".", 1)[0]

    rendered = f"{abs(scaled):,.{decimals}f}" if grouped else f"{abs(scaled):.{decimals}f}"
    width = len(rendered) + prefix_suffix
    if "%" in core:
        width += 1
    if scaled < 0:
        width += 1
    return width


DEFAULT_COL_WIDTH = 8.43


def detect_col_overflow(path: Path) -> dict:
    """Excel renders ######## when a *formatted* number is wider than its
    column. Byte-level assertions see a correct number format and a correct
    value; the reader sees a column of hashes.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(path))
    offenders = []
    undecidable = 0
    for ws in wb.worksheets:
        default_w = getattr(ws.sheet_format, "defaultColWidth", None) or DEFAULT_COL_WIDTH
        for row in ws.iter_rows():
            for cell in row:
                fmt = cell.number_format or "General"
                if fmt == "General" or cell.value is None:
                    continue
                need = _format_width(cell.value, fmt)
                if need is None:
                    undecidable += 1
                    continue
                dim = ws.column_dimensions.get(cell.column_letter)
                width = dim.width if (dim and dim.width) else default_w
                if need > width + 0.5:  # half a char of slack for padding
                    offenders.append(
                        {
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "format": fmt,
                            "needs_chars": round(need, 1),
                            "column_width": round(width, 1),
                        }
                    )
    present = True if offenders else (None if undecidable else False)
    return _result(
        "col_overflow",
        present,
        offending_cells=len(offenders),
        undecidable_cells=undecidable,
        samples=offenders[:5],
    )


# ==========================================================================
# VT3 — Word table wider than the text column
# ==========================================================================
def detect_table_overflow(path: Path) -> dict:
    """Sum of the table grid vs the section's usable text width."""
    doc = _docx_part(path, "word/document.xml")
    if doc is None:
        return _result("table_overflow", None, reason="no word/document.xml")

    sect = doc.find(f".//{W}sectPr")
    if sect is None:
        return _result("table_overflow", None, reason="no sectPr")
    pg = sect.find(f"{W}pgSz")
    mar = sect.find(f"{W}pgMar")
    page_w = int(pg.get(f"{W}w", 12240)) if pg is not None else 12240
    left = int(mar.get(f"{W}left", 1440)) if mar is not None else 1440
    right = int(mar.get(f"{W}right", 1440)) if mar is not None else 1440
    text_w = page_w - left - right

    offenders = []
    for i, tbl in enumerate(doc.iter(f"{W}tbl")):
        grid = tbl.find(f"{W}tblGrid")
        if grid is None:
            continue
        cols = [int(c.get(f"{W}w", 0)) for c in grid.findall(f"{W}gridCol")]
        # tblGrid is only a hint; Word lays a fixed-layout table out from the
        # per-cell tcW, and the two disagree whenever an agent sets cell widths
        # without touching the grid. Take the widest declared tcW per column
        # and fall back to the grid for columns that declare none.
        widest: dict[int, int] = {}
        for row in tbl.findall(f"{W}tr"):
            for ci, tc in enumerate(row.findall(f"{W}tc")):
                tcw = tc.find(f"{W}tcPr/{W}tcW")
                if tcw is None or tcw.get(f"{W}type") != "dxa":
                    continue
                w = int(tcw.get(f"{W}w", 0))
                span = tc.find(f"{W}tcPr/{W}gridSpan")
                if span is not None and int(span.get(f"{W}val", 1)) > 1:
                    continue  # merged cells don't map 1:1 to grid columns
                widest[ci] = max(widest.get(ci, 0), w)
        effective = [
            widest.get(ci, cols[ci] if ci < len(cols) else 0)
            for ci in range(max(len(cols), *(widest.keys() or [0]), 0) or len(cols))
        ]
        total = sum(effective)
        if not total:
            continue
        cols = effective
        indent_el = tbl.find(f"{W}tblPr/{W}tblInd")
        indent = int(indent_el.get(f"{W}w", 0)) if indent_el is not None else 0
        if total + indent > text_w:
            offenders.append(
                {
                    "table_index": i,
                    "grid_twips": total,
                    "indent_twips": indent,
                    "text_width_twips": text_w,
                    "overflow_twips": total + indent - text_w,
                    "columns": cols,
                }
            )
    return _result(
        "table_overflow",
        bool(offenders),
        text_width_twips=text_w,
        offending_tables=len(offenders),
        samples=offenders[:5],
    )


# ==========================================================================
# VT4 — chart anchored over its own data
# ==========================================================================
def _resolve_target(owner_part: str, target: str) -> str:
    """OPC relationship targets come in three flavours and openpyxl, Excel and
    officecli each pick a different one: absolute ("/xl/drawings/d1.xml"),
    parent-relative ("../drawings/d1.xml") and sibling-relative
    ("worksheets/sheet1.xml"). Resolve all three against the owning part.
    """
    import posixpath

    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(owner_part), target))


def _sheet_drawing_map(z: zipfile.ZipFile) -> dict:
    """sheet part name -> drawing part name."""
    out = {}
    for name in z.namelist():
        if not re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
            continue
        rels = f"xl/worksheets/_rels/{Path(name).name}.rels"
        if rels not in z.namelist():
            continue
        for rel in ET.fromstring(z.read(rels)).findall(f"{REL}Relationship"):
            if rel.get("Type", "").endswith("/drawing"):
                out[name] = _resolve_target(name, rel.get("Target", ""))
    return out


def _sheet_order(z: zipfile.ZipFile) -> list[tuple[str, str]]:
    """(sheet title, sheetN.xml part name) in workbook order."""
    wb = ET.fromstring(z.read("xl/workbook.xml"))
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    rels = {
        r.get("Id"): _resolve_target("xl/workbook.xml", r.get("Target", ""))
        for r in ET.fromstring(z.read("xl/_rels/workbook.xml.rels")).findall(f"{REL}Relationship")
    }
    return [
        (sh.get("name"), rels.get(sh.get(f"{ORel}id"), ""))
        for sh in wb.findall(f"{ns}sheets/{ns}sheet")
    ]


def detect_chart_overlap(path: Path) -> dict:
    """A chart whose anchor rectangle intersects the used range covers the very
    numbers it is plotting. openpyxl drops charts on read, so this reads the
    drawing XML directly.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(path))
    used = {}
    for ws in wb.worksheets:
        if ws.max_row and ws.max_column:
            used[ws.title] = (ws.min_row, ws.min_column, ws.max_row, ws.max_column)

    offenders = []
    charts = 0
    with zipfile.ZipFile(path) as z:
        drawings = _sheet_drawing_map(z)
        titles = dict(( part, name) for name, part in _sheet_order(z))
        for sheet_part, drawing_part in drawings.items():
            if drawing_part not in z.namelist():
                continue
            title = titles.get(sheet_part)
            root = ET.fromstring(z.read(drawing_part))
            for anchor in list(root.findall(f"{XDR}twoCellAnchor")) + list(
                root.findall(f"{XDR}oneCellAnchor")
            ):
                if anchor.find(f".//{XDR}graphicFrame") is None:
                    continue
                frm = anchor.find(f"{XDR}from")
                to = anchor.find(f"{XDR}to")
                if frm is None:
                    continue
                charts += 1
                c0 = int(frm.findtext(f"{XDR}col", "0")) + 1
                r0 = int(frm.findtext(f"{XDR}row", "0")) + 1
                if to is not None:
                    c1 = int(to.findtext(f"{XDR}col", "0")) + 1
                    r1 = int(to.findtext(f"{XDR}row", "0")) + 1
                else:  # oneCellAnchor: assume a typical 8x15 cell footprint
                    c1, r1 = c0 + 7, r0 + 14
                bounds = used.get(title)
                if not bounds:
                    continue
                ur0, uc0, ur1, uc1 = bounds
                rows_hit = not (r1 < ur0 or r0 > ur1)
                cols_hit = not (c1 < uc0 or c0 > uc1)
                if rows_hit and cols_hit:
                    offenders.append(
                        {
                            "sheet": title,
                            "chart_range": f"r{r0}c{c0}:r{r1}c{c1}",
                            "used_range": f"r{ur0}c{uc0}:r{ur1}c{uc1}",
                        }
                    )
    return _result(
        "chart_overlap",
        bool(offenders) if charts else False,
        charts_found=charts,
        offending_charts=len(offenders),
        samples=offenders[:5],
    )


# ==========================================================================
# VT5 — invisible text (contrast)
# ==========================================================================
def _srgb_channel(c: float) -> float:
    c = c / 255.0
    return c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4


def _luminance(rgb: str) -> float | None:
    if not rgb or not isinstance(rgb, str):
        return None
    hexpart = rgb[-6:]
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", hexpart):
        return None
    r, g, b = (int(hexpart[i : i + 2], 16) for i in (0, 2, 4))
    return 0.2126 * _srgb_channel(r) + 0.7152 * _srgb_channel(g) + 0.0722 * _srgb_channel(b)


def contrast_ratio(fg: str, bg: str) -> float | None:
    lf, lb = _luminance(fg), _luminance(bg)
    if lf is None or lb is None:
        return None
    hi, lo = max(lf, lb), min(lf, lb)
    return round((hi + 0.05) / (lo + 0.05), 2)


# Below this the text is not "low contrast", it is functionally invisible.
INVISIBLE_RATIO = 1.6


def detect_contrast(path: Path, threshold: float = INVISIBLE_RATIO) -> dict:
    """Text whose color has almost no contrast against its own cell fill.

    Every byte is valid and every assertion about "font color is white" passes.
    The cell simply reads as empty.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(path))
    offenders = []
    undecidable = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None or not str(cell.value).strip():
                    continue
                fill = cell.fill
                if not fill or fill.patternType != "solid":
                    continue
                bg = getattr(fill.fgColor, "rgb", None)
                fg = getattr(cell.font.color, "rgb", None) if cell.font and cell.font.color else None
                if not isinstance(bg, str) or not isinstance(fg, str):
                    undecidable += 1
                    continue
                ratio = contrast_ratio(fg, bg)
                if ratio is None:
                    undecidable += 1
                    continue
                if ratio < threshold:
                    offenders.append(
                        {
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "text": str(cell.value)[:30],
                            "fg": fg,
                            "bg": bg,
                            "contrast": ratio,
                        }
                    )
    present = True if offenders else (None if undecidable else False)
    return _result(
        "contrast",
        present,
        offending_cells=len(offenders),
        undecidable_cells=undecidable,
        samples=offenders[:5],
    )


# ==========================================================================
# VT6 — orphaned heading (render-only; no structural proxy exists)
# ==========================================================================
def detect_orphan_heading(pdf_path: Path, heading: str) -> dict:
    """Is `heading` the last thing on its page, with its section body pushed
    onto the next one?

    There is deliberately no OOXML shortcut here. Whether a heading strands at
    a page foot is decided by Word's layout engine at render time and by
    nothing in the file. This is the trap that exists to be unreachable
    without actually looking at the rendered page.
    """
    import fitz

    doc = fitz.open(str(pdf_path))
    try:
        target = heading.strip()
        for pno in range(doc.page_count):
            page = doc[pno]
            blocks = [b for b in page.get_text("blocks") if (b[4] or "").strip()]
            if not blocks:
                continue
            blocks.sort(key=lambda b: b[1])  # by top edge
            # Exact match only. A substring match happily finds the heading's
            # own words inside a body sentence and reports the wrong block.
            hit = next(
                (i for i, b in enumerate(blocks) if (b[4] or "").strip() == target), None
            )
            if hit is None:
                continue
            is_last = hit == len(blocks) - 1
            has_more_pages = any(
                (t := doc[p].get_text().strip()) for p in range(pno + 1, doc.page_count)
            )
            return _result(
                "orphan_heading",
                bool(is_last and has_more_pages),
                heading_page=pno + 1,
                blocks_on_page=len(blocks),
                heading_block_index=hit,
                heading_is_last_block=is_last,
                pages=doc.page_count,
            )
        return _result("orphan_heading", None, reason=f"heading {heading!r} not found in render")
    finally:
        doc.close()


# ==========================================================================
# VT7 — wrapped text clipped by a fixed row height
# ==========================================================================
DEFAULT_ROW_HEIGHT = 15.0  # points, one line at the default font size


def _display_width(text: str) -> int:
    """Excel column width counts CJK glyphs as two character cells."""
    return sum(2 if CJK.search(ch) else 1 for ch in text)


def detect_row_clip(path: Path) -> dict:
    """wrap_text turned on while the row keeps a one-line height: Excel shows
    only the first line and silently hides the rest. The value is intact, the
    alignment is exactly what was asked for, and the content is invisible.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(path))
    offenders = []
    for ws in wb.worksheets:
        default_w = getattr(ws.sheet_format, "defaultColWidth", None) or DEFAULT_COL_WIDTH
        for row in ws.iter_rows():
            for cell in row:
                al = cell.alignment
                if not al or not al.wrap_text or cell.value is None:
                    continue
                text = str(cell.value)
                if not text.strip():
                    continue
                dim = ws.column_dimensions.get(cell.column_letter)
                width = dim.width if (dim and dim.width) else default_w
                if width <= 0:
                    continue
                lines = max(1, -(-_display_width(text) // int(width)))
                rd = ws.row_dimensions.get(cell.row)
                height = rd.height if (rd and rd.height) else None
                if height is None:
                    continue  # auto-fit: Excel grows the row, nothing is hidden
                needed = lines * DEFAULT_ROW_HEIGHT
                if height + 1.0 < needed:
                    offenders.append(
                        {
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "wrapped_lines": lines,
                            "row_height": round(height, 1),
                            "needed_height": round(needed, 1),
                            "text": text[:30],
                        }
                    )
    return _result(
        "row_clip",
        bool(offenders),
        offending_cells=len(offenders),
        samples=offenders[:5],
    )


# ==========================================================================
# VT8 — column labels destroyed by a spanning merge
# ==========================================================================
def detect_header_lost(path: Path, labels: list | None = None) -> dict:
    """Did the column labels survive?

    Merging a header row to make a spanning title keeps only the top-left
    value; every other label is discarded. Nothing about the file is invalid —
    the table just no longer says what its columns mean.
    """
    from openpyxl import load_workbook

    labels = labels or []
    if not labels:
        return _result("header_lost", None, reason="no expected labels supplied")
    wb = load_workbook(str(path))
    present = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                for lab in labels:
                    if text == lab:
                        present.add(lab)
    missing = [lab for lab in labels if lab not in present]
    return _result(
        "header_lost",
        bool(missing),
        expected=labels,
        surviving=sorted(present),
        missing=missing,
    )


# ==========================================================================
# VT9 — text drawn on top of text (render-only)
# ==========================================================================
def detect_text_collision(pdf_path: Path, min_overlap: float = 24.0) -> dict:
    """Two text blocks whose rendered boxes genuinely overlap.

    A footer placed into a bottom margin that is too small is drawn straight
    through the body text. Like the orphan heading, this is decided by the
    layout engine at render time — the file contains a valid margin and a valid
    footer, and nothing that says they collide.
    """
    import fitz

    doc = fitz.open(str(pdf_path))
    try:
        hits = []
        for pno in range(doc.page_count):
            blocks = [
                b for b in doc[pno].get_text("blocks") if (b[4] or "").strip()
            ]
            for i in range(len(blocks)):
                for j in range(i + 1, len(blocks)):
                    ax0, ay0, ax1, ay1 = blocks[i][:4]
                    bx0, by0, bx1, by1 = blocks[j][:4]
                    ox = min(ax1, bx1) - max(ax0, bx0)
                    oy = min(ay1, by1) - max(ay0, by0)
                    if ox <= 0 or oy <= 0:
                        continue
                    area = ox * oy
                    if area < min_overlap:
                        continue
                    hits.append(
                        {
                            "page": pno + 1,
                            "overlap_area": round(area, 1),
                            "a": (blocks[i][4] or "").strip()[:34],
                            "b": (blocks[j][4] or "").strip()[:34],
                        }
                    )
        return _result(
            "text_collision",
            bool(hits),
            pages=doc.page_count,
            collisions=len(hits),
            samples=hits[:5],
        )
    finally:
        doc.close()


def detect_print_split(pdf_path: Path, labels: list | None = None) -> dict:
    """Are the columns of one table printed onto more than one page?

    Widening a column past the printable width makes Excel spill the remaining
    columns onto a separate sheet of paper. On screen everything looks fine —
    the tear only exists in the laid-out artefact, so only the render can see
    it. Every column width in the file is individually valid.
    """
    import fitz

    labels = labels or []
    if not labels:
        return _result("print_split", None, reason="no expected labels supplied")
    doc = fitz.open(str(pdf_path))
    try:
        where = {}
        for pno in range(doc.page_count):
            text = doc[pno].get_text()
            for lab in labels:
                if lab in text:
                    where.setdefault(lab, []).append(pno + 1)
        pages_used = sorted({p for ps in where.values() for p in ps})
        found = [lab for lab in labels if lab in where]
        if len(found) < 2:
            return _result(
                "print_split", None,
                reason="fewer than two header labels found in the render",
                found=found, pages=doc.page_count,
            )
        return _result(
            "print_split",
            len(pages_used) > 1,
            pages=doc.page_count,
            header_pages={k: v for k, v in where.items()},
            missing=[lab for lab in labels if lab not in where],
        )
    finally:
        doc.close()


# ==========================================================================
# Dispatch
# ==========================================================================
STRUCTURAL = {
    "cjk_font_slot": detect_cjk_slot,
    "col_overflow": detect_col_overflow,
    "table_overflow": detect_table_overflow,
    "chart_overlap": detect_chart_overlap,
    "contrast": detect_contrast,
    "row_clip": detect_row_clip,
    "header_lost": detect_header_lost,
}

RENDER_ONLY = {"orphan_heading", "print_split"}


def run_detector(trap: str, output: Path | None, pdf: Path | None = None, **kw) -> dict:
    if output is None or not Path(output).exists():
        return _result(trap, None, reason="no output file")
    if trap in STRUCTURAL:
        try:
            return STRUCTURAL[trap](Path(output), **kw)
        except Exception as exc:  # a detector crash must never read as "clean"
            return _result(trap, None, reason=f"detector error: {exc!r}"[:200])
    if trap in RENDER_ONLY:
        if pdf is None or not Path(pdf).exists():
            return _result(trap, None, reason="render required but missing")
        try:
            if trap == "orphan_heading":
                return detect_orphan_heading(Path(pdf), kw.get("heading", ""))
            return detect_print_split(Path(pdf), kw.get("labels"))
        except Exception as exc:
            return _result(trap, None, reason=f"detector error: {exc!r}"[:200])
    return _result(trap, None, reason=f"unknown trap {trap}")
