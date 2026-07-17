import pytest
from openpyxl import Workbook

from office_agent.adapters.excel_adapter import ExcelAdapter
from office_agent.schemas.operations import FillMode, StyleParams


@pytest.fixture
def ws():
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Score"
    sheet["A2"] = "Alice"
    sheet["B2"] = 90
    return sheet


def test_read_cells_range(ws):
    result = ExcelAdapter.read_cells(ws, (1, 1, 2, 2))
    assert result["data"] == [["Name", "Score"], ["Alice", 90]]


def test_write_cells_overwrite(ws):
    result = ExcelAdapter.write_cells(ws, (3, 1, None, None), [["Bob", 85]])
    assert ws["A3"].value == "Bob"
    assert result["affected"] == ["A3", "B3"]


def test_write_cells_merge_keeps_existing(ws):
    result = ExcelAdapter.write_cells(
        ws, (1, 1, None, None), [["X", "Y"]], FillMode.MERGE
    )
    assert ws["A1"].value == "Name"  # kept
    assert result["affected"] == []  # nothing actually changed


def test_write_cells_append_concatenates(ws):
    ExcelAdapter.write_cells(ws, (1, 1, None, None), [["!"]], FillMode.APPEND)
    assert ws["A1"].value == "Name!"


def test_write_cells_empty_values_raises(ws):
    with pytest.raises(ValueError, match="non-empty"):
        ExcelAdapter.write_cells(ws, None, [])


def test_edit_formula_requires_equals(ws):
    with pytest.raises(ValueError, match="start with"):
        ExcelAdapter.edit_formula(ws, (3, 3, 3, 3), "SUM(B2:B3)")


def test_edit_formula(ws):
    result = ExcelAdapter.edit_formula(ws, (3, 3, 3, 3), "=SUM(B2:B2)")
    assert ws["C3"].value == "=SUM(B2:B2)"
    assert result["affected"] == ["C3"]


def test_apply_style(ws):
    ExcelAdapter.apply_style(
        ws, (1, 1, 1, 2), StyleParams(bold=True, bg_color="#FFFF00")
    )
    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.start_color.rgb.endswith("FFFF00")


def test_apply_style_requires_range(ws):
    with pytest.raises(ValueError, match="range is required"):
        ExcelAdapter.apply_style(ws, None, StyleParams(bold=True))


def test_insert_rows_executes_once(ws):
    ExcelAdapter.insert_rows_cols(ws, 2, 1, "row")
    assert ws["A2"].value is None  # new empty row
    assert ws["A3"].value == "Alice"  # shifted down exactly once


def test_insert_bad_type_raises(ws):
    with pytest.raises(ValueError, match="'row' or 'column'"):
        ExcelAdapter.insert_rows_cols(ws, 1, 1, "diagonal")


def test_create_chart(ws):
    result = ExcelAdapter.create_chart(ws, "A1:B2", "bar", {"title": "T"})
    assert result["chart_type"] == "bar"
    assert len(ws._charts) == 1  # exactly one chart


def test_create_chart_bad_range(ws):
    with pytest.raises(ValueError, match="Invalid chart data range"):
        ExcelAdapter.create_chart(ws, "!!!", "bar")


def test_create_chart_bad_type(ws):
    with pytest.raises(ValueError, match="Unsupported chart type"):
        ExcelAdapter.create_chart(ws, "A1:B2", "scatter3d")


def test_conditional_select_contains(ws):
    matches = ExcelAdapter.conditional_select(ws, {"contains": "Alice"})
    assert matches == ["A2"]


def test_conditional_select_value_gt(ws):
    matches = ExcelAdapter.conditional_select(ws, {"value_gt": 50})
    assert matches == ["B2"]


def test_conditional_select_unknown_key_raises(ws):
    with pytest.raises(ValueError, match="Unknown condition keys"):
        ExcelAdapter.conditional_select(ws, {"colour": "red"})
