"""bench/BUGS.md OA-4: an abnormal termination must never hand back nothing.

Budget exhaustion, user abort, and giving up on verification all now persist
the working copy to a NEW path beside the original.
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from fake_llm import FakeLLM, FakeMessage, text_block, tool_use

from office_agent.agent.loop import AgentSession, BaseUI
from office_agent.config import Budgets, Config


class RecordingUI(BaseUI):
    def __init__(self):
        self.notices = []

    def notify(self, message):
        self.notices.append(message)

    def emit_text(self, text):
        pass


def _session(script, **budget_kwargs):
    config = Config(
        api_key="fake",
        visual_verify=False,
        budgets=Budgets(**budget_kwargs) if budget_kwargs else Budgets(),
    )
    ui = RecordingUI()
    return AgentSession(config, ui=ui, llm=FakeLLM(script=script)), ui


def test_budget_exhaustion_rescues_the_working_copy(excel_doc_path):
    """E5 in the evaluation burned 30 calls and produced None. Now the partial
    work lands on disk and the model is told where."""
    script = [
        FakeMessage(content=[tool_use("open_document", {"file_path": str(excel_doc_path)})])
    ]
    for i in range(4):
        script.append(
            FakeMessage(
                content=[
                    tool_use(
                        "excel_write_cells",
                        {
                            "doc_id": "PLACEHOLDER",
                            "sheet": None,
                            "range": f"H{i + 1}",
                            "values": [[f"partial-{i}"]],
                        },
                    )
                ]
            )
        )
    script.append(FakeMessage(content=[text_block("budget gone")]))

    session, ui = _session(script, max_tool_calls_per_task=3)
    # Resolve doc_id/sheet lazily: patch the scripted inputs after open.
    original_handle = session._handle_tool
    state = {}

    def handle(name, tool_input):
        if tool_input.get("doc_id") == "PLACEHOLDER":
            tool_input["doc_id"] = state["doc_id"]
            tool_input["sheet"] = state["sheet"]
        out = original_handle(name, tool_input)
        if name == "open_document":
            payload = json.loads(out[0])
            state["doc_id"] = payload["doc_id"]
            state["sheet"] = payload["summary"]["sheets"][0]
        return out

    session._handle_tool = handle
    result = session.run_turn("do a lot")

    rescued = [p for p in result.saved_paths if ".partial-" in p]
    assert rescued, f"nothing rescued; saved_paths={result.saved_paths}"
    out = Path(rescued[0])
    assert out.exists() and out.parent == excel_doc_path.parent
    # The original is untouched…
    assert load_workbook(str(excel_doc_path))[state["sheet"]]["H1"].value is None
    # …and the rescue copy carries the partial work.
    assert load_workbook(str(out))[state["sheet"]]["H1"].value == "partial-0"
    assert any("已把部分成果另存为" in n for n in ui.notices)


def test_rescue_path_is_reported_to_the_model(excel_doc_path):
    """The budget-exhausted tool result must name the rescue path so the
    model's summary can tell the user where its work went."""
    script = [
        FakeMessage(content=[tool_use("open_document", {"file_path": str(excel_doc_path)})]),
        FakeMessage(content=[tool_use("get_structure", {"doc_id": "X"})]),
        FakeMessage(content=[text_block("done")]),
    ]
    session, _ = _session(script, max_tool_calls_per_task=1)
    original_handle = session._handle_tool
    state = {}

    def handle(name, tool_input):
        if tool_input.get("doc_id") == "X":
            tool_input["doc_id"] = state["doc_id"]
        out = original_handle(name, tool_input)
        if name == "open_document":
            state["doc_id"] = json.loads(out[0])["doc_id"]
        return out

    session._handle_tool = handle
    session.run_turn("go")
    # Only open_document ran (no edits), so there is nothing to rescue — but
    # the contract field must still be present and honest.
    payloads = [
        json.loads(c["messages"][-1]["content"][0]["content"])
        for c in session.llm.calls
        if isinstance(c["messages"][-1]["content"], list)
        and c["messages"][-1]["content"][0].get("type") == "tool_result"
    ]
    exhausted = [p for p in payloads if "budget" in str(p.get("error", ""))]
    assert exhausted and exhausted[0]["rescued_paths"] == []


def test_no_rescue_when_the_model_already_saved(excel_doc_path):
    script = [
        FakeMessage(content=[tool_use("open_document", {"file_path": str(excel_doc_path)})]),
        FakeMessage(content=[tool_use("save_document", {"doc_id": "X"})]),
        FakeMessage(content=[text_block("saved")]),
    ]
    session, _ = _session(script)
    original_handle = session._handle_tool
    state = {}

    def handle(name, tool_input):
        if tool_input.get("doc_id") == "X":
            tool_input["doc_id"] = state["doc_id"]
        out = original_handle(name, tool_input)
        if name == "open_document":
            state["doc_id"] = json.loads(out[0])["doc_id"]
        return out

    session._handle_tool = handle
    session.run_turn("save it")
    session.abort_requested = True
    before = list(session.saved_paths)
    assert session._rescue_unsaved("test") == []
    assert session.saved_paths == before


def test_rescue_is_idempotent(excel_doc_path):
    """Repeated terminations must not litter the directory with copies."""
    script = [
        FakeMessage(content=[tool_use("open_document", {"file_path": str(excel_doc_path)})]),
        FakeMessage(content=[text_block("ok")]),
    ]
    session, _ = _session(script)
    session.run_turn("open it")
    doc_id = next(iter(session.ctx.sessions.sessions))
    sheet = session.ctx.sessions.get(doc_id).doc.sheetnames[0]
    # Force an edit so there is something worth rescuing.
    from office_agent.tools import REGISTRY

    REGISTRY.dispatch(
        session.ctx,
        "excel_write_cells",
        {"doc_id": doc_id, "sheet": sheet, "range": "H1", "values": [["x"]]},
    )
    first = session._rescue_unsaved("first")
    second = session._rescue_unsaved("second")
    assert len(first) == 1 and second == []
    assert len(list(excel_doc_path.parent.glob("*.partial-*"))) == 1
