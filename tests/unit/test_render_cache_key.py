"""The render cache must key on document CONTENT, not on file bytes.

python-docx and openpyxl rebuild the zip on every save, stamping each member
with the current time (DOS timestamps: 2-second granularity), and openpyxl also
rewrites dcterms:modified in docProps/core.xml. A raw byte hash therefore
changes when nothing changed, and the harness re-exports through Word/Excel for
nothing — every visual verification pays a full AppleScript round trip.

These tests are offline: the export functions are stubbed.
"""
import shutil
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from office_agent.core.session import EditSession
from office_agent.render import renderer as R
from office_agent.render.renderer import Renderer, content_digest


@pytest.fixture
def fake_export(monkeypatch):
    """Replace the Word/Excel round trip with a counter + a stub PDF."""
    calls = []

    def stub(src, pdf_path, sheet=None, timeout=120.0):
        calls.append(str(src))
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        pdf_path.write_bytes(b"%PDF-1.4 stub")
        return pdf_path

    monkeypatch.setattr(R, "export_docx_to_pdf", stub)
    monkeypatch.setattr(R, "export_xlsx_to_pdf", stub)
    monkeypatch.setattr(
        R, "pdf_to_images", lambda pdf, out_dir: [R.PageImage(0, pdf, 100, 100)]
    )
    return calls


def _bump_zip_timestamps(path: Path, date_time=(2030, 1, 1, 0, 0, 0)) -> None:
    """Rewrite the package with different member mtimes, identical content —
    exactly what a second save of unchanged content produces."""
    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(path) as src, zipfile.ZipFile(
        tmp, "w", zipfile.ZIP_DEFLATED
    ) as dst:
        for item in src.infolist():
            info = zipfile.ZipInfo(item.filename, date_time=date_time)
            info.compress_type = item.compress_type
            dst.writestr(info, src.read(item.filename))
    tmp.replace(path)


def test_digest_ignores_zip_member_timestamps(word_doc_path):
    before = content_digest(word_doc_path)
    _bump_zip_timestamps(word_doc_path)
    assert content_digest(word_doc_path) == before


def test_digest_ignores_the_modified_timestamp(excel_doc_path, tmp_path):
    """docProps/core.xml is rewritten on every openpyxl save and never renders."""
    before = content_digest(excel_doc_path)
    other = tmp_path / "other.xlsx"
    shutil.copy2(excel_doc_path, other)
    with zipfile.ZipFile(excel_doc_path) as z:
        core = z.read("docProps/core.xml").decode()
    assert "dcterms:modified" in core

    tampered = tmp_path / "tampered.xlsx"
    with zipfile.ZipFile(excel_doc_path) as src, zipfile.ZipFile(
        tampered, "w", zipfile.ZIP_DEFLATED
    ) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "docProps/core.xml":
                data = data.replace(b"2026", b"2099")
            dst.writestr(item, data)
    assert content_digest(tampered) == before


def test_digest_still_catches_a_real_edit(excel_doc_path):
    before = content_digest(excel_doc_path)
    wb = load_workbook(str(excel_doc_path))
    wb[wb.sheetnames[0]]["Z99"] = "edited"
    wb.save(str(excel_doc_path))
    assert content_digest(excel_doc_path) != before


def test_digest_falls_back_on_a_non_package(tmp_path):
    junk = tmp_path / "not-a-zip.docx"
    junk.write_bytes(b"definitely not a zip")
    first = content_digest(junk)  # returns a hash instead of raising
    assert first
    junk.write_bytes(b"different bytes")
    # Fallback must stay content-sensitive: a constant would cache a stale
    # render forever.
    assert content_digest(junk) != first


def test_repeated_render_of_unchanged_content_is_one_export(word_doc_path, fake_export):
    """The regression this fixes: flush() restamps the zip, so the second
    render used to miss the cache and pay a full Word export."""
    session = EditSession(str(word_doc_path))
    try:
        r = Renderer(session)
        r.render()
        _bump_zip_timestamps(session.working_path)  # simulate a later save
        r.render()
        r.render()
        assert len(fake_export) == 1
    finally:
        session.cleanup()


def test_an_edit_between_renders_still_re_exports(excel_doc_path, fake_export):
    session = EditSession(str(excel_doc_path))
    try:
        r = Renderer(session)
        r.render()
        session.doc[session.doc.sheetnames[0]]["Z99"] = "changed"
        session.dirty = True
        r.render()
        assert len(fake_export) == 2
    finally:
        session.cleanup()


def test_sheet_selection_is_part_of_the_key(excel_doc_path, fake_export):
    session = EditSession(str(excel_doc_path))
    try:
        r = Renderer(session)
        first = session.doc.sheetnames[0]
        r.render(sheet=first)
        r.render(sheet=first)
        assert len(fake_export) == 1
        r.render(sheet=None)
        assert len(fake_export) == 2  # different view of the same bytes
    finally:
        session.cleanup()
