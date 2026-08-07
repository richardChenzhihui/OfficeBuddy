"""OA-1 minimal repro — see ../BUGS.md.

Isolate which openpyxl load flag makes a re-saved xlsx unopenable by real Excel.

Expected output on the buggy code (openpyxl 3.1.5, macOS Excel):
    A_untouched      render=OK
    B_default        render=OK
    C_keepvba        render=FAIL ... 参数错误。 (-50)
    D_richtext       render=OK
    E_both           render=FAIL ... (-50)
    F_editsession    render=FAIL ... (-50)
i.e. keep_vba=True alone is sufficient to break the file; rich_text is innocent.

Run:  python bench/repro/xlsx_keepvba_corruption.py


Variants, all starting from the same pristine fixture bytes:
  A  untouched          (control — must render)
  B  default flags      load_workbook(p)                      -> save
  C  keep_vba           load_workbook(p, keep_vba=True)       -> save
  D  rich_text          load_workbook(p, rich_text=True)      -> save
  E  keep_vba+rich_text (== EditSession._load)                -> save
  F  EditSession(...).flush()

Each variant is staged into the Excel sandbox container and exported to PDF.
Excel is quit between variants so a -50 failure cannot cascade.
"""
import shutil
import subprocess
import sys
import time
from pathlib import Path

BENCH = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BENCH))

from openpyxl import load_workbook  # noqa: E402
from office_agent.core.session import EXCEL_CONTAINER, EditSession  # noqa: E402
from office_agent.render.applescript import export_xlsx_to_pdf  # noqa: E402

SCRATCH = Path(__file__).resolve().parent / "_work"
SCRATCH.mkdir(exist_ok=True)
SRC = SCRATCH / "pristine.xlsx"


def build_fixture():
    from fixtures import sales_book
    sales_book(SRC)
    return SRC


def quit_excel():
    subprocess.run(
        ["osascript", "-e", 'tell application "Microsoft Excel" to quit saving no'],
        capture_output=True, text=True, timeout=60,
    )
    time.sleep(2)


def variant(name, fn):
    work = SCRATCH / f"v_{name}"
    shutil.rmtree(work, ignore_errors=True)
    work.mkdir(parents=True)
    p = work / "book.xlsx"
    shutil.copy(SRC, p)
    fn(p)
    return p


def zip_diff(a: Path, b: Path):
    import zipfile
    za = {n: zipfile.ZipFile(a).getinfo(n).file_size for n in zipfile.ZipFile(a).namelist()}
    zb = {n: zipfile.ZipFile(b).getinfo(n).file_size for n in zipfile.ZipFile(b).namelist()}
    lost = sorted(set(za) - set(zb))
    added = sorted(set(zb) - set(za))
    return lost, added


def render(p: Path, tag: str) -> str:
    container = EXCEL_CONTAINER / "repro_flush" / tag
    container.mkdir(parents=True, exist_ok=True)
    staged = container / "book.xlsx"
    shutil.copy(p, staged)
    pdf = container / "out.pdf"
    try:
        export_xlsx_to_pdf(staged, pdf, timeout=120.0)
        return "OK"
    except Exception as exc:
        return f"FAIL {type(exc).__name__}: {str(exc)[:160]}"
    finally:
        shutil.rmtree(container, ignore_errors=True)


VARIANTS = {
    "A_untouched": lambda p: None,
    "B_default": lambda p: load_workbook(p).save(p),
    "C_keepvba": lambda p: load_workbook(p, keep_vba=True).save(p),
    "D_richtext": lambda p: load_workbook(p, rich_text=True).save(p),
    "E_both": lambda p: load_workbook(p, keep_vba=True, keep_links=True, rich_text=True).save(p),
    "F_editsession": lambda p: shutil.copy(EditSession(str(p)).flush(), p),
}

if __name__ == "__main__":
    build_fixture()
    print(f"fixture: {SRC} ({SRC.stat().st_size} bytes)\n")
    for name, fn in VARIANTS.items():
        try:
            p = variant(name, fn)
        except Exception as exc:
            print(f"{name:16} BUILD-FAIL {exc!r}")
            continue
        lost, added = zip_diff(SRC, p)
        quit_excel()
        res = render(p, name)
        print(f"{name:16} size={p.stat().st_size:>7}  render={res}")
        if lost:
            print(f"{'':16}   lost parts: {lost}")
        if added:
            print(f"{'':16}   added parts: {added}")
    quit_excel()
