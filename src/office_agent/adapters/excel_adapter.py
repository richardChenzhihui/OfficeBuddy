"""Excel adapter wrapping openpyxl. Stateless: mutates the worksheet passed in."""
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.chart import BarChart, LineChart, PieChart
from openpyxl.chart.reference import Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries

from ..schemas.operations import FillMode, StyleParams

Coords = Tuple[int, int, Optional[int], Optional[int]]  # min_row, min_col, max_row, max_col


class ExcelAdapter:
    """Stateless operations on openpyxl worksheets."""

    @staticmethod
    def read_cells(
        worksheet: Any, coords: Optional[Coords], include_formula: bool = False
    ) -> Dict[str, Any]:
        from openpyxl.cell.rich_text import CellRichText

        def _plain(value):
            # rich_text=True loads keep intra-cell formatting; tools return
            # the concatenated plain text (JSON-safe) rather than the object.
            return str(value) if isinstance(value, CellRichText) else value

        data: List[List[Any]] = []
        formulas: List[List[Any]] = []
        if coords is None:
            rows = worksheet.iter_rows()
        else:
            min_row, min_col, max_row, max_col = coords
            rows = worksheet.iter_rows(
                min_row=min_row,
                min_col=min_col,
                max_row=max_row or worksheet.max_row,
                max_col=max_col or worksheet.max_column,
            )
        for row in rows:
            data.append([_plain(cell.value) for cell in row])
            formulas.append(
                [cell.value if cell.data_type == "f" else None for cell in row]
            )
        return {"data": data, "formulas": formulas if include_formula else None}

    @staticmethod
    def write_cells(
        worksheet: Any,
        coords: Optional[Coords],
        values: List[List[Any]],
        fill_mode: FillMode = FillMode.OVERWRITE,
    ) -> Dict[str, Any]:
        if not values:
            raise ValueError("values must be a non-empty list of row lists.")
        if coords is None:
            min_row, min_col = 1, 1
        else:
            min_row, min_col = coords[0], coords[1]

        affected: List[str] = []
        for i, row_values in enumerate(values):
            for j, value in enumerate(row_values):
                cell = worksheet.cell(row=min_row + i, column=min_col + j)
                if fill_mode == FillMode.OVERWRITE:
                    cell.value = value
                elif fill_mode == FillMode.APPEND:
                    cell.value = value if cell.value is None else str(cell.value) + str(value)
                elif fill_mode == FillMode.MERGE:
                    if cell.value is None:
                        cell.value = value
                    else:
                        continue  # existing value kept; not counted as affected
                affected.append(f"{get_column_letter(min_col + j)}{min_row + i}")
        return {"affected": affected}

    @staticmethod
    def edit_formula(
        worksheet: Any, coords: Optional[Coords], formula: str
    ) -> Dict[str, Any]:
        if coords is None:
            raise ValueError("A cell range is required for formula editing.")
        if not formula.startswith("="):
            raise ValueError(f"Formula must start with '=': got '{formula}'.")
        min_row, min_col, max_row, max_col = coords
        affected: List[str] = []
        for row_idx in range(min_row, (max_row or min_row) + 1):
            for col_idx in range(min_col, (max_col or min_col) + 1):
                worksheet.cell(row=row_idx, column=col_idx).value = formula
                affected.append(f"{get_column_letter(col_idx)}{row_idx}")
        return {"affected": affected, "formula": formula}

    @staticmethod
    def apply_style(
        worksheet: Any, coords: Optional[Coords], style: StyleParams
    ) -> Dict[str, Any]:
        if coords is None:
            raise ValueError("A cell range is required for style editing.")
        min_row, min_col, max_row, max_col = coords

        font_kwargs: Dict[str, Any] = {}
        if style.font_name:
            font_kwargs["name"] = style.font_name
        if style.font_size:
            font_kwargs["size"] = style.font_size
        if style.bold is not None:
            font_kwargs["bold"] = style.bold
        if style.italic is not None:
            font_kwargs["italic"] = style.italic
        if style.underline is not None:
            font_kwargs["underline"] = "single" if style.underline else None
        if style.color:
            font_kwargs["color"] = style.color.lstrip("#")
        font = Font(**font_kwargs) if font_kwargs else None

        fill = None
        if style.bg_color:
            bg = style.bg_color.lstrip("#")
            fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        alignment = Alignment(horizontal=style.alignment) if style.alignment else None

        affected: List[str] = []
        for row_idx in range(min_row, (max_row or worksheet.max_row) + 1):
            for col_idx in range(min_col, (max_col or worksheet.max_column) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
                if style.number_format:
                    cell.number_format = style.number_format
                affected.append(f"{get_column_letter(col_idx)}{row_idx}")
        return {"affected": affected}

    @staticmethod
    def insert_rows_cols(
        worksheet: Any, position: int, count: int, insert_type: str = "row"
    ) -> Dict[str, Any]:
        if insert_type not in ("row", "column"):
            raise ValueError(
                f"insert_type must be 'row' or 'column', got '{insert_type}'."
            )
        if position < 1 or count < 1:
            raise ValueError("position and count must be >= 1 (Excel is 1-indexed).")
        if insert_type == "row":
            worksheet.insert_rows(position, count)
            affected = [f"rows {position}..{position + count - 1}"]
        else:
            worksheet.insert_cols(position, count)
            affected = [
                f"columns {get_column_letter(position)}.."
                f"{get_column_letter(position + count - 1)}"
            ]
        return {"affected": affected, "type": insert_type, "position": position, "count": count}

    @staticmethod
    def freeze_panes(worksheet: Any, cell: Optional[str]) -> Dict[str, Any]:
        """Freeze rows above / columns left of `cell` ('A2' = freeze row 1).
        cell=None (or 'A1') unfreezes."""
        if cell is None or str(cell).strip() == "":
            worksheet.freeze_panes = None
            return {"freeze_panes": None, "affected": [f"{worksheet.title}: panes unfrozen"]}
        cell = str(cell).strip().upper().replace("$", "")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell)
        except Exception as exc:
            raise ValueError(
                f"Invalid freeze cell '{cell}': {exc}. Use a single A1-style cell "
                "such as 'A2' (freeze the header row) or 'B2' (freeze row 1 and column A)."
            ) from exc
        if (min_col, min_row) != (max_col, max_row):
            raise ValueError(
                f"freeze_panes takes ONE cell, not a range: got '{cell}'. "
                "'A2' freezes row 1; 'B2' freezes row 1 and column A."
            )
        worksheet.freeze_panes = cell
        if cell == "A1":
            return {"freeze_panes": None, "affected": [f"{worksheet.title}: panes unfrozen"]}
        frozen = []
        if min_row > 1:
            frozen.append(f"rows 1..{min_row - 1}")
        if min_col > 1:
            frozen.append(f"columns A..{get_column_letter(min_col - 1)}")
        return {
            "freeze_panes": cell,
            "affected": [f"{worksheet.title}: froze {' and '.join(frozen)}"],
        }

    # Excel's own worksheet-name rules; violating them makes the file unopenable.
    _INVALID_SHEET_CHARS = set(r"[]:*?/\\")

    @classmethod
    def _validate_sheet_name(cls, name: str, workbook: Any) -> str:
        name = (name or "").strip()
        if not name:
            raise ValueError("Sheet name must be a non-empty string.")
        if len(name) > 31:
            raise ValueError(
                f"Sheet name '{name}' is {len(name)} chars: Excel allows at most 31."
            )
        bad = sorted(set(name) & cls._INVALID_SHEET_CHARS)
        if bad:
            raise ValueError(
                f"Sheet name '{name}' contains characters Excel forbids: {bad}."
            )
        existing = {s.lower() for s in workbook.sheetnames}
        if name.lower() in existing:
            raise ValueError(
                f"A sheet named '{name}' already exists (sheets: {workbook.sheetnames}). "
                "Pick another name, or write into the existing sheet."
            )
        return name

    @classmethod
    def create_sheet(
        cls, workbook: Any, name: str, index: Optional[int] = None
    ) -> Dict[str, Any]:
        name = cls._validate_sheet_name(name, workbook)
        worksheet = workbook.create_sheet(title=name, index=index)
        return {
            "sheet": worksheet.title,
            "index": workbook.sheetnames.index(worksheet.title),
            "sheets": list(workbook.sheetnames),
            "affected": [f"created sheet '{worksheet.title}'"],
        }

    @staticmethod
    def delete_sheet(workbook: Any, name: str) -> Dict[str, Any]:
        if name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{name}' not found: workbook has sheets {workbook.sheetnames}."
            )
        if len(workbook.sheetnames) == 1:
            raise ValueError(
                f"Refusing to delete '{name}': a workbook must keep at least one sheet."
            )
        del workbook[name]
        return {
            "sheets": list(workbook.sheetnames),
            "affected": [f"deleted sheet '{name}'"],
        }

    @classmethod
    def rename_sheet(cls, workbook: Any, name: str, new_name: str) -> Dict[str, Any]:
        if name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{name}' not found: workbook has sheets {workbook.sheetnames}."
            )
        if new_name == name:
            raise ValueError(f"new_name is identical to the current name '{name}'.")
        new_name = cls._validate_sheet_name(new_name, workbook)
        workbook[name].title = new_name
        return {
            "sheet": new_name,
            "sheets": list(workbook.sheetnames),
            "affected": [f"renamed sheet '{name}' -> '{new_name}'"],
            "note": (
                "Formulas in OTHER sheets that referenced the old name are NOT "
                "rewritten by this engine — re-check any cross-sheet formulas."
            ),
        }

    @classmethod
    def copy_sheet(
        cls, workbook: Any, name: str, new_name: Optional[str] = None
    ) -> Dict[str, Any]:
        if name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{name}' not found: workbook has sheets {workbook.sheetnames}."
            )
        copy = workbook.copy_worksheet(workbook[name])
        if new_name:
            copy.title = cls._validate_sheet_name(new_name, workbook)
        return {
            "sheet": copy.title,
            "sheets": list(workbook.sheetnames),
            "affected": [f"copied sheet '{name}' -> '{copy.title}'"],
            "note": (
                "openpyxl's sheet copy carries values, formulas and styles but "
                "NOT charts, images or data validations."
            ),
        }

    @staticmethod
    def _default_chart_anchor(worksheet: Any, data_max_col: Optional[int]) -> str:
        """Right of everything already on the sheet, so the chart never sits on
        top of the data (OA-5). Two columns of breathing room."""
        rightmost = max(worksheet.max_column or 1, data_max_col or 1)
        return f"{get_column_letter(rightmost + 2)}1"

    @staticmethod
    def create_chart(
        worksheet: Any,
        data_range: str,
        chart_type: str,
        options: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        options = options or {}
        try:
            min_col, min_row, max_col, max_row = range_boundaries(data_range)
        except Exception as exc:
            raise ValueError(
                f"Invalid chart data range '{data_range}': {exc}. Use A1-style like 'A1:B10'."
            ) from exc

        chart_classes = {"bar": BarChart, "line": LineChart, "pie": PieChart}
        if chart_type not in chart_classes:
            raise ValueError(
                f"Unsupported chart type '{chart_type}': expected one of {sorted(chart_classes)}."
            )
        chart = chart_classes[chart_type]()

        titles = options.get("titles_from_data", True)
        # Conventional layout: first column holds category labels. Without
        # set_categories they'd become a bogus extra data series.
        use_categories = (
            options.get("first_column_as_categories", True)
            and max_col is not None
            and max_col > min_col
        )
        if use_categories:
            data = Reference(
                worksheet,
                min_col=min_col + 1,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            )
            categories = Reference(
                worksheet,
                min_col=min_col,
                min_row=min_row + (1 if titles else 0),
                max_row=max_row,
            )
            chart.add_data(data, titles_from_data=titles)
            chart.set_categories(categories)
        else:
            data = Reference(
                worksheet,
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            )
            chart.add_data(data, titles_from_data=titles)
        if "title" in options:
            chart.title = options["title"]
        anchor = options.get("chart_cell") or ExcelAdapter._default_chart_anchor(
            worksheet, max_col
        )
        worksheet.add_chart(chart, anchor)
        return {
            "chart_type": chart_type,
            "data_range": data_range,
            "chart_cell": anchor,
            "affected": [f"chart@{anchor}"],
        }

    @staticmethod
    def delete_chart(worksheet: Any, chart_index: Optional[int] = None) -> Dict[str, Any]:
        """Delete one chart by index, or all charts when chart_index is None."""
        charts = getattr(worksheet, "_charts", [])
        if not charts:
            raise ValueError("This sheet has no charts to delete.")
        if chart_index is None:
            removed = len(charts)
            worksheet._charts = []
            return {"affected": [f"removed all {removed} chart(s)"]}
        if not 0 <= chart_index < len(charts):
            raise ValueError(
                f"chart_index {chart_index} out of range: sheet has "
                f"{len(charts)} chart(s) (0..{len(charts)-1})."
            )
        del worksheet._charts[chart_index]
        return {"affected": [f"removed chart {chart_index}"]}

    @staticmethod
    def conditional_select(worksheet: Any, condition: Dict[str, Any]) -> List[str]:
        known_keys = {"contains", "value_gt", "has_formula"}
        unknown = set(condition) - known_keys
        if unknown:
            raise ValueError(
                f"Unknown condition keys {sorted(unknown)}: supported keys are {sorted(known_keys)}."
            )
        matches: List[str] = []
        for row in worksheet.iter_rows():
            for cell in row:
                match = True
                if "contains" in condition:
                    if cell.value is None or condition["contains"] not in str(cell.value):
                        match = False
                if match and "value_gt" in condition:
                    try:
                        if cell.value is None or float(cell.value) <= condition["value_gt"]:
                            match = False
                    except (ValueError, TypeError):
                        match = False
                if match and "has_formula" in condition:
                    is_formula = cell.data_type == "f"
                    if bool(condition["has_formula"]) != is_formula:
                        match = False
                if match:
                    matches.append(f"{get_column_letter(cell.column)}{cell.row}")
        return matches
