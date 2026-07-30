import os
import tempfile
from pathlib import Path

import pytest

# Unit tests must never touch the Office app sandbox containers: on machines
# without the TCC container-access grant the syscalls HANG (not error). Render
# tests (mac_office marker) override this per-test where container placement
# actually matters.
os.environ.setdefault(
    "OFFICE_AGENT_WORK_ROOT",
    str(Path(tempfile.gettempdir()) / "office_agent_test_work"),
)


@pytest.fixture
def word_doc_path(tmp_path):
    """Minimal .docx fixture: a single Normal-styled paragraph, no run formatting."""
    import docx

    dst = tmp_path / "test.docx"
    doc = docx.Document()
    doc.add_paragraph("测试段落")
    doc.save(dst)
    return dst


@pytest.fixture
def excel_doc_path(tmp_path):
    """Minimal .xlsx fixture: one sheet with a two-column header row."""
    import openpyxl

    dst = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Age"
    wb.save(dst)
    return dst
